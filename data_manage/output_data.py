from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font


def update_excel(path: str, big_dist_data: dict, mid_dist_data: dict):
    """recives a list of lists with shape:
    [integral_value, start_integral, end_integral]
    and returns and creates an excel file with the data in the
    output path. Creates output directory if it doesn't exist."""

    # Create workbook
    wb = Workbook()
    wb = load_workbook(path, read_only=False)
    ws = wb.worksheets[0]

    ##General metrics##
    for i, key in enumerate(["slope start-finish", "upper area", "downer area"]):
        ws.cell(row=6, column=6 + i).value = key
        ws.cell(row=6, column=6 + i).fill = PatternFill(start_color="75E6DA", end_color="75E6DA", fill_type = "solid")
        ws.cell(row=7, column=6 + i).value = big_dist_data[key]
        big_dist_data.pop(key)

    # Write first set of data
    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "Bigger Distance Pivot"
    title_cell.fill = PatternFill(start_color="FF4A4A", end_color="FF4A4A", fill_type = "solid")
    title_cellfont = Font(size=12)
    for i, (name, value) in enumerate(big_dist_data.items()):
        ws.cell(row=1, column=i+6).value = name + " BDP"
        ws.cell(row=1, column=i+6).fill = PatternFill(start_color="75E6DA", end_color="75E6DA", fill_type = "solid")
        ws.cell(row=2, column=i+6).value = value

    

    # Write second set
    title_cell = ws.cell(row=3, column=5)
    title_cell.value = "Middle Distance Pivot"
    title_cell.fill = PatternFill(start_color="FF4A4A", end_color="FF4A4A", fill_type = "solid")
    #Copied and pasted from excel file

    for i, (name, value) in enumerate(mid_dist_data.items()):
        ws.cell(row=3, column=i+6).value = name + " MDP"
        ws.cell(row=3, column=i+6).fill = PatternFill(start_color="75E6DA", end_color="75E6DA", fill_type = "solid")
        ws.cell(row=4, column=i+6).value = value

    for col in ws.columns:
        column = col[0].column_letter   # Get the column name
        ws.column_dimensions[column].width = 20
    # Save excel file
    wb.save(path)


