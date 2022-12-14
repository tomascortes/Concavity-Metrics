import string
import os
import sys
from openpyxl import Workbook, load_workbook

class MoreValuesThanTimeStamps(Exception):
    """Raised when there are missing timestamps """
    pass


def path_selector() -> string:
    # Declare the dictionry options
    options = dict()
    counter = 1

    # Print all the excel files in directory
    files = os.listdir("./")
    for file in files:
        if ".xlsx" in file:
            options[counter] = file
            print(f"{counter})", end="")
            print(file)
            counter += 1
    # Ask for input
    answer = input("Write number of the file to be selected an press enter \n")
    while not answer.isdigit():
        print("\nRemember to write just a number")
        answer = input("Write number of file to be selected an press enter \n")
    answer_number = int(answer)
    return options[answer_number]


def read_file(path: str, work_sheet: str=None) -> tuple:
    """recives a path of a excel file and return a tuple of two lists, 
    corresponding to time and Values"""
    wb = Workbook()
    wb = load_workbook(path, read_only=True, data_only=True)
    if work_sheet:
        ws = wb[work_sheet]
    else:
        ws = wb.worksheets[0]

    output_time = []
    output_values = []

    for row in ws.rows:
        output_time.append(row[0].value)
        output_values.append(row[1].value)

    start_x = round(ws.cell(2, 3).value, 2)
    finish_x = round(ws.cell(2, 4).value, 2)


    output_time.pop(0)
    output_values.pop(0)

    # Sometimes the data read Nones at the end,
    # there should be no Nones in the middle

    for i in range(len(output_time)):
        if output_time[i] == None:
            output_time = output_time[:i]
            break

    for i in range(len(output_values)):
        if output_values[i] == None:
            output_values = output_values[:i]
            break
    if len(output_values) > len(output_time):
        raise MoreValuesThanTimeStamps("There are more input values than time values")
    elif len(output_values) != len(output_time):
        output_time = output_time[:len(output_values)]
    return (output_time, output_values, start_x, finish_x)

def get_verbose() -> bool:
    n = len(sys.argv)
    if n >= 2:
        return sys.argv[1] == "p"
    return False


